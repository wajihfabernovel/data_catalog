"""Import and export helpers for the user journey mapping module."""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


JOURNEY_TEMPLATE_HEADERS = [
    "Journey ID",
    "Journey Name",
    "Module/Domain",
    "Step #",
    "User Action",
    "Screen/Component",
    "Tables Read (comma-separated)",
    "Tables Written (comma-separated)",
    "Status Field Changes",
    "Validation Rules Applied",
    "Business Rules/Logic",
    "Notes",
]

JOURNEY_INDEX_HEADERS = [
    "Journey ID",
    "Journey Name",
    "Module/Domain",
    "Frequency (Daily/Weekly/Monthly/Ad-hoc)",
    "Complexity (Low/Med/High)",
    "Total Steps",
    "Core Tables Involved",
    "Interviewer",
]

TABLE_CROSS_REFERENCE_HEADERS = [
    "Table Name",
    "Domain",
    "Journey IDs (comma-separated)",
    "Read Count",
    "Write Count",
    "Access Pattern",
    "Centrality Score",
    "Legacy Table Type",
    "Target Entity (proposed)",
]

INSTRUCTIONS_LINES = [
    "USER JOURNEY CAPTURE TEMPLATE — INSTRUCTIONS",
    "",
    "PURPOSE",
    "This template captures user journeys to bridge legacy table inventory → new data model design.",
    "Each journey documents the sequence of user actions, the data touched at each step, and the business rules enforced.",
    "",
    "HOW TO USE THIS TEMPLATE",
    "",
    "1. JOURNEY INDEX SHEET",
    "   - Start here for each new user journey",
    "   - Assign a unique Journey ID (J001, J002, etc.)",
    "   - Document high-level journey metadata: name, domain, frequency, complexity",
    "",
    "2. JOURNEY TEMPLATE SHEET",
    "   - Add one row per step in the journey",
    "   - Keep step numbers sequential within each journey",
    "   - Use comma-separated table names for reads and writes",
    "",
    "3. TABLE CROSS-REFERENCE SHEET",
    "   - Review which tables are touched across journeys",
    "   - Use this to identify central and peripheral tables",
    "",
    "4. NEXT STEPS",
    "   - Validate cross-reference findings",
    "   - Define target entities and ownership",
]


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _normalize_csv_list(raw_value: Any) -> list[str]:
    text = _safe_text(raw_value)
    if not text or text.upper() == "NONE":
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def parse_write_tables(raw_value: Any) -> list[dict[str, str]]:
    writes: list[dict[str, str]] = []
    for item in _normalize_csv_list(raw_value):
        if item.endswith(")") and " (" in item:
            table_name, operation = item.rsplit(" (", 1)
            writes.append(
                {"table_name": table_name.strip(), "write_operation": operation[:-1].strip().upper()}
            )
        else:
            writes.append({"table_name": item, "write_operation": "UPDATE"})
    return writes


def format_transition_summary(transitions: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for transition in transitions:
        entity = _safe_text(transition.get("entity_table"))
        field_name = _safe_text(transition.get("status_field_name"))
        from_state = _safe_text(transition.get("from_state")) or "NULL"
        to_state = _safe_text(transition.get("to_state"))
        label = ".".join(part for part in [entity, field_name] if part)
        if label and to_state:
            parts.append(f"{label}: {from_state} → {to_state}")
    return ", ".join(parts) if parts else "None"


def build_journey_workbook(
    journeys: list[dict[str, Any]],
    steps_by_journey: dict[str, list[dict[str, Any]]],
    analysis_rows: list[dict[str, Any]],
    state_transitions: list[dict[str, Any]],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    ws_template = workbook.create_sheet("Journey Template")
    ws_template.append(JOURNEY_TEMPLATE_HEADERS)
    for cell in ws_template[1]:
        cell.font = Font(bold=True)

    ws_index = workbook.create_sheet("Journey Index")
    ws_index.append(JOURNEY_INDEX_HEADERS)
    for cell in ws_index[1]:
        cell.font = Font(bold=True)

    ws_cross = workbook.create_sheet("Table Cross-Reference")
    ws_cross.append(TABLE_CROSS_REFERENCE_HEADERS)
    for cell in ws_cross[1]:
        cell.font = Font(bold=True)

    ws_instructions = workbook.create_sheet("Instructions")
    for idx, line in enumerate(INSTRUCTIONS_LINES, start=1):
        ws_instructions.cell(row=idx, column=1, value=line)
    ws_instructions["A1"].font = Font(bold=True)

    for journey in journeys:
        journey_id = journey["journey_id"]
        steps = sorted(steps_by_journey.get(journey_id, []), key=lambda item: item["step_number"])
        core_tables = sorted(
            {
                table["table_name"]
                for step in steps
                for table in step.get("table_refs", [])
                if table.get("table_name")
            }
        )
        ws_index.append(
            [
                journey_id,
                journey.get("journey_name", ""),
                journey.get("module_domain", ""),
                journey.get("frequency", ""),
                journey.get("complexity", ""),
                len(steps),
                ", ".join(core_tables),
                journey.get("interviewer", ""),
            ]
        )
        for step in steps:
            read_tables = [
                ref["table_name"]
                for ref in step.get("table_refs", [])
                if ref.get("access_mode") == "READ"
            ]
            write_tables = []
            for ref in step.get("table_refs", []):
                if ref.get("access_mode") != "WRITE":
                    continue
                op = _safe_text(ref.get("write_operation")) or "UPDATE"
                write_tables.append(f"{ref['table_name']} ({op})")
            ws_template.append(
                [
                    journey_id,
                    journey.get("journey_name", ""),
                    journey.get("module_domain", ""),
                    step.get("step_number", ""),
                    step.get("user_action", ""),
                    step.get("screen_component", ""),
                    ", ".join(read_tables) if read_tables else "None",
                    ", ".join(write_tables) if write_tables else "None",
                    step.get("status_field_changes", "") or "None",
                    step.get("validation_rules", "") or "None",
                    step.get("business_rules", "") or "None",
                    step.get("notes", "") or "None",
                ]
            )

    for row in analysis_rows:
        ws_cross.append(
            [
                row.get("table_name", ""),
                row.get("domain", ""),
                row.get("journey_ids", ""),
                row.get("read_count", 0),
                row.get("write_count", 0),
                row.get("access_pattern", ""),
                row.get("centrality_score", ""),
                row.get("legacy_table_type", ""),
                row.get("target_entity_proposed", ""),
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_journey_workbook(file_bytes: bytes) -> list[dict[str, Any]]:
    sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, dtype=str)
    index_df = sheets.get("Journey Index", pd.DataFrame()).fillna("")
    template_df = sheets.get("Journey Template", pd.DataFrame()).fillna("")

    journeys: dict[str, dict[str, Any]] = {}
    for _, row in index_df.iterrows():
        journey_id = _safe_text(row.get("Journey ID"))
        if not journey_id:
            continue
        journeys[journey_id] = {
            "journey": {
                "journey_id": journey_id,
                "journey_name": _safe_text(row.get("Journey Name")),
                "module_domain": _safe_text(row.get("Module/Domain")),
                "primary_user_role": _safe_text(row.get("Primary User Role")),
                "frequency": _safe_text(row.get("Frequency (Daily/Weekly/Monthly/Ad-hoc)")),
                "complexity": _safe_text(row.get("Complexity (Low/Med/High)")),
                "interview_date": _safe_text(row.get("Interview Date")),
                "interviewer": _safe_text(row.get("Interviewer")),
                "scrum_team": _safe_text(row.get("Scrum Team")),
            },
            "steps": [],
            "step_tables": [],
            "transitions": [],
        }

    for _, row in template_df.iterrows():
        journey_id = _safe_text(row.get("Journey ID"))
        if not journey_id:
            continue
        bucket = journeys.setdefault(
            journey_id,
            {
                "journey": {
                    "journey_id": journey_id,
                    "journey_name": _safe_text(row.get("Journey Name")),
                    "module_domain": _safe_text(row.get("Module/Domain")),
                    "primary_user_role": "",
                    "frequency": "",
                    "complexity": "",
                    "interview_date": "",
                    "interviewer": "",
                    "scrum_team": "",
                },
                "steps": [],
                "step_tables": [],
                "transitions": [],
            },
        )
        step_number = int(_safe_text(row.get("Step #")) or "0")
        bucket["steps"].append(
            {
                "step_number": step_number,
                "user_action": _safe_text(row.get("User Action")),
                "screen_component": _safe_text(row.get("Screen/Component")),
                "status_field_changes": _safe_text(row.get("Status Field Changes")),
                "validation_rules": _safe_text(row.get("Validation Rules Applied")),
                "business_rules": _safe_text(row.get("Business Rules/Logic")),
                "notes": _safe_text(row.get("Notes")),
            }
        )
        for table_name in _normalize_csv_list(row.get("Tables Read (comma-separated)")):
            bucket["step_tables"].append(
                {
                    "journey_id": journey_id,
                    "step_number": step_number,
                    "table_name": table_name,
                    "access_mode": "READ",
                    "write_operation": None,
                }
            )
        for write in parse_write_tables(row.get("Tables Written (comma-separated)")):
            bucket["step_tables"].append(
                {
                    "journey_id": journey_id,
                    "step_number": step_number,
                    "table_name": write["table_name"],
                    "access_mode": "WRITE",
                    "write_operation": write["write_operation"],
                }
            )

    return list(journeys.values())
