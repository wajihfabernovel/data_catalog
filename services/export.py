"""Excel export service for the data catalog workbook."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils.helpers import sanitize_sheet_name, serialize_date


SECTION_FILL = PatternFill("solid", fgColor="2F75B5")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=13)
LABEL_FILL = PatternFill("solid", fgColor="DDEBF7")
SIGNOFF_FILL = PatternFill("solid", fgColor="E2F0D9")
VALUE_FILL = PatternFill("solid", fgColor="FFF2CC")
BODY_FILL = PatternFill("solid", fgColor="F7F7F7")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def _style_section_header(ws, row: int, title: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 5):
        ws.cell(row=row, column=column).border = THIN_BORDER


def _style_field_row(ws, row: int, label: str, helper_text: str, value: str, label_fill=None) -> None:
    label_fill = label_fill or LABEL_FILL
    cells = [
        ws.cell(row=row, column=1, value=label),
        ws.cell(row=row, column=2, value=helper_text),
        ws.cell(row=row, column=3, value=value),
        ws.cell(row=row, column=4, value=""),
    ]

    cells[0].fill = label_fill
    cells[0].font = Font(bold=True, color="1F3B6D")

    cells[1].fill = BODY_FILL
    cells[1].font = Font(italic=True, color="5A5A5A")

    cells[2].fill = VALUE_FILL
    cells[3].fill = VALUE_FILL

    for cell in cells:
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_schema_table(ws, start_row: int, schema: list[dict], table_name: str) -> int:
    _style_section_header(ws, start_row, "1. SCHEMA")
    header_row = start_row + 1
    headers = ["column_name", "edm_type", "sql_type"]
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=index, value=header)
        cell.fill = LABEL_FILL
        cell.font = Font(bold=True, color="1F3B6D")
        cell.border = THIN_BORDER

    data_start = header_row + 1
    for row_index, column in enumerate(schema, start=data_start):
        ws.cell(row=row_index, column=1, value=column["column_name"])
        ws.cell(row=row_index, column=2, value=column["edm_type"])
        ws.cell(row=row_index, column=3, value=column["sql_type"])
        for column_index in range(1, 4):
            ws.cell(row=row_index, column=column_index).border = THIN_BORDER

    last_row = max(data_start, data_start + len(schema) - 1)
    table_ref = f"A{header_row}:C{last_row}"
    excel_table = Table(
        displayName=f"{sanitize_sheet_name(table_name).replace(' ', '_')[:20]}_schema",
        ref=table_ref,
    )
    excel_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(excel_table)
    return last_row + 2


def _write_table_summary(ws, table: dict) -> int:
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = table["table_name"]
    title.font = Font(size=16, bold=True, color="1F1F1F")
    title.alignment = Alignment(horizontal="left")

    _style_field_row(
        ws,
        2,
        "Primary key",
        "Primary key parsed from Dataverse metadata",
        table.get("primary_key", ""),
    )
    _style_field_row(
        ws,
        3,
        "Column count",
        "Number of parsed property columns",
        str(len(table.get("schema", []))),
    )
    return 5


def _write_data_quality(ws, start_row: int, table: dict) -> int:
    _style_section_header(ws, start_row, "3. DATA QUALITY OBSERVATIONS")
    data = table["data_quality"]
    rows = [
        ("Nullable issues", "Observed nullability problems", data["nullable_issues"]),
        ("Format inconsistencies", "Observed format inconsistencies", data["format_inconsistencies"]),
        ("Duplicate records", "YES / NO", data["duplicate_records"]),
        ("Orphan records", "YES / NO", data["orphan_records"]),
        ("Hard delete in use", "YES / NO", data["hard_delete_in_use"]),
        ("Overall quality rating", "CLEAN / ACCEPTABLE / PROBLEMATIC", data["overall_quality_rating"]),
        ("Quality notes", "Additional quality comments", data["quality_notes"]),
    ]
    row = start_row + 1
    for label, helper_text, value in rows:
        _style_field_row(ws, row, label, helper_text, value)
        row += 1
    return row + 1


def _write_pipeline(ws, start_row: int, table: dict) -> int:
    _style_section_header(ws, start_row, "4. PIPELINE RELEVANCE")
    data = table["pipeline"]
    rows = [
        ("Extract by pipeline?", "YES / NO / UNSURE - should the data platform extract this?", data["extract_by_pipeline"]),
        ("Delta extraction column", "Column that drives incremental load", data["delta_extraction_column"]),
        ("Feed Power BI?", "YES / NO / UNSURE", data["feed_power_bi"]),
        ("Key metrics or dimensions", "What does this table contribute to analytics?", data["key_metrics_or_dimensions"]),
        ("Write path", "STORED PROCEDURE / DAB / DIRECT SQL / UNKNOWN", data["write_path"]),
    ]
    row = start_row + 1
    for label, helper_text, value in rows:
        _style_field_row(ws, row, label, helper_text, value)
        row += 1
    return row + 1


def _write_target_model(ws, start_row: int, table: dict) -> int:
    _style_section_header(ws, start_row, "5. TARGET MODEL RECOMMENDATION")
    data = table["target_model"]
    rows = [
        ("Recommendation", "KEEP AS IS / REFACTOR / MERGE WITH / SPLIT INTO / RETIRE", data["recommendation"]),
        ("If merge - merge with", "Which table should this be merged with?", data["merge_with"]),
        ("If split - split into", "Which tables should this be split into?", data["split_into"]),
        ("If retire - replaced by", "What replaces this table?", data["replaced_by"]),
        ("Missing columns needed", "Columns needed in target that do not exist today", data["missing_columns"]),
        ("Missing constraints needed", "Constraints that should be added", data["missing_constraints"]),
    ]
    row = start_row + 1
    for label, helper_text, value in rows:
        _style_field_row(ws, row, label, helper_text, value)
        row += 1
    return row + 1


def _write_signoff(ws, start_row: int, table: dict) -> int:
    _style_section_header(ws, start_row, "6. SIGN OFF")
    data = table["signoff"]
    rows = [
        ("Completed by", "Name of person who filled in this template", data["completed_by"]),
        ("Reviewed by", "Name of data lead who reviewed", data["reviewed_by"]),
        ("Reviewed by (business)", "Name of business stakeholder who validated", data["reviewed_by_business"]),
        ("Status", "DRAFT / IN REVIEW / APPROVED", data["status"]),
        ("Date approved", "Date approved", serialize_date(data["date_approved"]) or ""),
        ("Notes", "Any additional comments or follow up actions", data["notes"]),
    ]
    row = start_row + 1
    for label, helper_text, value in rows:
        _style_field_row(ws, row, label, helper_text, value, label_fill=SIGNOFF_FILL)
        row += 1
    return row + 1


def _write_relationships(ws, start_row: int, table: dict) -> int:
    relationships = table.get("relationships", {})
    _style_section_header(ws, start_row, "2. RELATIONSHIPS")

    row = start_row + 1
    reference_headers = [
        "FK column in this table",
        "References table",
        "References column",
        "Cardinality",
        "Mandatory?",
    ]
    for column_index, header in enumerate(reference_headers, start=1):
        cell = ws.cell(row=row, column=column_index, value=header)
        cell.fill = LABEL_FILL
        cell.font = Font(bold=True, color="1F3B6D")
        cell.border = THIN_BORDER

    reference_rows = relationships.get("references", []) or [{}]
    for entry in reference_rows:
        row += 1
        values = [
            entry.get("fk_column", ""),
            entry.get("references_table", ""),
            entry.get("references_column", ""),
            entry.get("cardinality", ""),
            entry.get("mandatory", ""),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    subheader = ws.cell(row=row, column=1, value="Tables that reference this table")
    subheader.fill = LABEL_FILL
    subheader.font = Font(bold=True, color="1F3B6D")
    subheader.border = THIN_BORDER

    row += 1
    referenced_by_headers = ["Table name", "Via column", "Cardinality"]
    for column_index, header in enumerate(referenced_by_headers, start=1):
        cell = ws.cell(row=row, column=column_index, value=header)
        cell.fill = LABEL_FILL
        cell.font = Font(bold=True, color="1F3B6D")
        cell.border = THIN_BORDER

    referenced_by_rows = relationships.get("referenced_by", []) or [{}]
    for entry in referenced_by_rows:
        row += 1
        values = [
            entry.get("table_name", ""),
            entry.get("via_column", ""),
            entry.get("cardinality", ""),
        ]
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=column_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return row + 2


def _apply_sheet_layout(ws) -> None:
    widths = {1: 30, 2: 52, 3: 36, 4: 18}
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width
    ws.freeze_panes = "A2"


def build_excel_workbook(tables: list[dict]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table in tables:
        sheet_name = sanitize_sheet_name(table["table_name"])
        worksheet = workbook.create_sheet(title=sheet_name)
        current_row = _write_table_summary(worksheet, table)
        current_row = _write_schema_table(worksheet, current_row, table.get("schema", []), table["table_name"])
        current_row = _write_relationships(worksheet, current_row, table)
        current_row = _write_data_quality(worksheet, current_row, table)
        current_row = _write_pipeline(worksheet, current_row, table)
        current_row = _write_target_model(worksheet, current_row, table)
        _write_signoff(worksheet, current_row, table)
        _apply_sheet_layout(worksheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
